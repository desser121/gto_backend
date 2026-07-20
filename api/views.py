from rest_framework import viewsets, status
from rest_framework.authentication import TokenAuthentication
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from django.http import HttpResponse
from django.contrib.auth.models import User
from django.db import transaction
from .models import (
    Normative, ParticipantList, Participant, Step, Exercise,
    TestResult, UserProfile, ROLE_CHOICES, ROLE_PERMISSIONS, AuditLog
)
from .serializers import (
    NormativeSerializer, ParticipantListSerializer,
    ParticipantListCreateUpdateSerializer, ParticipantSerializer,
    TestResultSerializer, StepSerializer, ExerciseSerializer
)
from .permissions import get_user_permissions, IsRoot, CanCreateList, CanDeleteList, CanEditList
from .audit import log_action
import openpyxl
from datetime import datetime
from io import BytesIO


class NormativeViewSet(viewsets.ModelViewSet):
    queryset = Normative.objects.all()
    serializer_class = NormativeSerializer

    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            return [IsAuthenticated()]
        return [IsAuthenticated()]

    def get_queryset(self):
        qs = super().get_queryset()
        step = self.request.query_params.get('step')
        exercise = self.request.query_params.get('exercise')
        if step:
            qs = qs.filter(step_id=step)
        if exercise:
            qs = qs.filter(exercise_id=exercise)
        return qs


class TestResultViewSet(viewsets.ModelViewSet):
    queryset = TestResult.objects.all()
    serializer_class = TestResultSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        participant_id = self.request.query_params.get('participant')
        if participant_id:
            queryset = queryset.filter(participant_id=participant_id)
        return queryset


class ParticipantListViewSet(viewsets.ModelViewSet):
    queryset = ParticipantList.objects.all()

    def get_permissions(self):
        if self.action == 'create':
            return [IsAuthenticated(), CanCreateList()]
        if self.action in ['destroy']:
            return [IsAuthenticated(), CanDeleteList()]
        if self.action in ['update', 'partial_update']:
            return [IsAuthenticated(), CanEditList()]
        return [IsAuthenticated()]

    def get_serializer_class(self):
        if self.action in ['list', 'retrieve']:
            return ParticipantListSerializer
        return ParticipantListCreateUpdateSerializer

    def create(self, request, *args, **kwargs):
        name = request.data.get('name', '').strip()
        if name and ParticipantList.objects.filter(name=name).exists():
            existing = ParticipantList.objects.get(name=name)
            return Response({
                'error': 'duplicate_name',
                'existing_id': existing.id,
                'existing_name': existing.name,
                'existing_count': existing.participants.count(),
            }, status=409)
        return super().create(request, *args, **kwargs)

    def perform_create(self, serializer):
        obj = serializer.save()
        log_action(self.request.user, 'create_list', 'Создан список участников', obj.name)

    def perform_destroy(self, instance):
        log_action(self.request.user, 'delete_list', 'Удалён список участников', instance.name)
        instance.delete()


class ParticipantViewSet(viewsets.ModelViewSet):
    queryset = Participant.objects.all()
    serializer_class = ParticipantSerializer

    def get_permissions(self):
        if self.action == 'create':
            return [IsAuthenticated(), CanEditList()]
        if self.action in ['destroy']:
            return [IsAuthenticated(), CanDeleteList()]
        if self.action in ['update', 'partial_update']:
            return [IsAuthenticated(), CanEditList()]
        return [IsAuthenticated()]

    def get_queryset(self):
        queryset = super().get_queryset()
        participant_list_id = self.request.query_params.get('participant_list')
        if participant_list_id:
            queryset = queryset.filter(participant_list_id=participant_list_id)
        return queryset


class StepViewSet(viewsets.ModelViewSet):
    queryset = Step.objects.all()
    serializer_class = StepSerializer

    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            return [IsAuthenticated()]
        return [IsAuthenticated()]


class ExerciseViewSet(viewsets.ModelViewSet):
    queryset = Exercise.objects.all()
    serializer_class = ExerciseSerializer

    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            return [IsAuthenticated()]
        return [IsAuthenticated()]


class CurrentUserView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        perms, role = get_user_permissions(user)
        return Response({
            "id": user.id,
            "username": user.username,
            "first_name": user.first_name,
            "last_name": user.last_name,
            "email": user.email,
            "role": role,
            "permissions": perms,
        })


class ExportFederalTemplateView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request):
        participants_data = request.data.get('participants', [])
        exercise_names = request.data.get('exercise_names', [])
        region = request.data.get('region', 'Удмуртская Республика')
        center_name = request.data.get('center_name', '')

        if not participants_data:
            return Response({"error": "Нет данных участников"}, status=400)
        if not exercise_names:
            return Response({"error": "Нет названий упражнений"}, status=400)

        template_path = 'api/federal_template.xlsx'
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        today = datetime.now().date()
        ws['D4'] = region
        if center_name:
            ws['D8'] = center_name
        if participants_data:
            first = participants_data[0]
            ws['F7'] = first.get('step', '')
            ws['G7'] = 'мужской' if first.get('gender', 'М') == 'М' else 'женский'
            ws['M7'] = f'<< {today.day} >>'
            ws['N7'] = self._month_name(today.month)
            ws['O7'] = f'{today.year} года'
        for i, name in enumerate(exercise_names):
            ws.cell(row=12, column=5 + i).value = name
        for idx, p in enumerate(participants_data):
            row = 13 + idx
            ws.cell(row=row, column=1).value = idx + 1
            ws.cell(row=row, column=2).value = p.get('fio', '')
            ws.cell(row=row, column=4).value = p.get('uin', '')
            ws.cell(row=row, column=3).value = ''
            values = p.get('values', {})
            for i, ex_name in enumerate(exercise_names):
                val = values.get(ex_name, '')
                ws.cell(row=row, column=5 + i).value = val if val else ''

        log_action(request.user, 'export_federal', 'Выгружен федеральный шаблон', f'{len(participants_data)} участников')

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        filename = f"federal_template_{today.strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def _calc_age(self, birthdate_str):
        if not birthdate_str:
            return 0
        try:
            bd = datetime.strptime(birthdate_str, '%Y-%m-%d').date()
            today = datetime.now().date()
            age = today.year - bd.year
            if (today.month, today.day) < (bd.month, bd.day):
                age -= 1
            return age
        except Exception:
            return 0

    def _month_name(self, m):
        months = {
            1: 'января', 2: 'февраля', 3: 'марта', 4: 'апреля',
            5: 'мая', 6: 'июня', 7: 'июля', 8: 'августа',
            9: 'сентября', 10: 'октября', 11: 'ноября', 12: 'декабря'
        }
        return months.get(m, '')


def _can_manage_users(user):
    _, role = get_user_permissions(user)
    return role in ('root', 'admin')


class UserManagementView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not _can_manage_users(request.user):
            return Response({'error': 'Нет прав'}, status=403)
        users = User.objects.all().select_related('profile')
        data = []
        for u in users:
            try:
                role = u.profile.role
            except UserProfile.DoesNotExist:
                role = 'viewer'
            data.append({
                'id': u.id,
                'username': u.username,
                'first_name': u.first_name,
                'last_name': u.last_name,
                'email': u.email,
                'is_active': u.is_active,
                'role': role,
            })
        return Response(data)

    def post(self, request):
        if not _can_manage_users(request.user):
            return Response({'error': 'Нет прав'}, status=403)

        _, my_role = get_user_permissions(request.user)
        username = request.data.get('username', '').strip()
        password = request.data.get('password', '')
        first_name = request.data.get('first_name', '')
        last_name = request.data.get('last_name', '')
        email = request.data.get('email', '')
        role = request.data.get('role', 'viewer')

        if not username or not password:
            return Response({'error': 'Логин и пароль обязательны'}, status=400)
        if User.objects.filter(username=username).exists():
            return Response({'error': 'Пользователь уже существует'}, status=400)
        valid_roles = [r[0] for r in ROLE_CHOICES]
        if role not in valid_roles:
            return Response({'error': 'Неверная роль'}, status=400)
        if my_role == 'admin' and role == 'root':
            return Response({'error': 'Администратор не может назначать роль суперадмина'}, status=403)

        user = User.objects.create_user(
            username=username, password=password,
            first_name=first_name, last_name=last_name, email=email
        )
        UserProfile.objects.create(user=user, role=role)
        log_action(request.user, 'create_user', f'Создан пользователь {username}', role)
        return Response({'id': user.id, 'username': user.username, 'role': role}, status=201)


class UserManagementDetailView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def put(self, request, user_id):
        if not _can_manage_users(request.user):
            return Response({'error': 'Нет прав'}, status=403)

        _, my_role = get_user_permissions(request.user)

        try:
            user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            return Response({'error': 'Пользователь не найден'}, status=404)

        role = request.data.get('role')
        if role:
            valid_roles = [r[0] for r in ROLE_CHOICES]
            if role not in valid_roles:
                return Response({'error': 'Неверная роль'}, status=400)
            if my_role == 'admin' and role == 'root':
                return Response({'error': 'Администратор не может назначать роль суперадмина'}, status=403)
            profile, _ = UserProfile.objects.get_or_create(user=user)
            old_role = profile.role
            profile.role = role
            profile.save()
            if old_role != role:
                log_action(request.user, 'change_role', f'Роль {user.username} изменена: {old_role} -> {role}', user.username)

        for field in ['first_name', 'last_name', 'email', 'is_active']:
            if field in request.data:
                setattr(user, field, request.data[field])

        new_password = request.data.get('password')
        if new_password:
            user.set_password(new_password)

        user.save()
        log_action(request.user, 'edit_user', f'Обновлён пользователь {user.username}', user.username)
        return Response({'ok': True})

    def delete(self, request, user_id):
        if not _can_manage_users(request.user):
            return Response({'error': 'Нет прав'}, status=403)

        try:
            user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            return Response({'error': 'Пользователь не найден'}, status=404)
        if user == request.user:
            return Response({'error': 'Нельзя удалить себя'}, status=400)
        log_action(request.user, 'delete_user', f'Удалён пользователь {user.username}', user.username)
        user.delete()
        return Response({'ok': True})


class RoleChoicesView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not _can_manage_users(request.user):
            return Response({'error': 'Нет прав'}, status=403)
        _, my_role = get_user_permissions(request.user)
        if my_role == 'admin':
            return Response([{'value': r[0], 'label': r[1]} for r in ROLE_CHOICES if r[0] != 'root'])
        return Response([{'value': r[0], 'label': r[1]} for r in ROLE_CHOICES])


class ActivityLogView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        _, role = get_user_permissions(user)
        limit = int(request.query_params.get('limit', 10))
        offset = int(request.query_params.get('offset', 0))
        if limit > 500:
            limit = 500

        qs = AuditLog.objects.select_related('user')
        if role not in ('root', 'admin'):
            qs = qs.filter(user=user)

        total = qs.count()
        logs = qs[offset:offset + limit]
        data = []
        for log in logs:
            data.append({
                'id': log.id,
                'username': log.user.username if log.user else 'system',
                'action': log.get_action_display(),
                'description': log.description,
                'target_name': log.target_name,
                'created_at': log.created_at.strftime('%d.%m.%Y %H:%M'),
            })
        return Response({'logs': data, 'total': total, 'offset': offset, 'limit': limit})
